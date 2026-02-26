"""
Import healers from Excel file into healers.json
"""
import json
import openpyxl
import re
from pathlib import Path

EXCEL_FILE = r"C:\Users\abise\Downloads\Healers Data (2).xlsx"
HEALERS_JSON = Path(__file__).parent.parent / "frontend" / "public" / "healers.json"


def clean_coord(val):
    """Parse coordinate from various formats like '15.1211° E', '74.0300° N', or plain numbers."""
    if val is None:
        return None
    s = str(val).strip()
    # Remove degree symbols and direction letters
    s = re.sub(r'[°\s]*[NSEW]$', '', s.strip())
    s = s.strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    print("Available sheets:", wb.sheetnames)
    
    # Try to find the best sheet - check all sheets for data
    all_healers = {}  # keyed by UID to deduplicate
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value).strip().lower() if c.value else '' for c in list(ws.iter_rows(1,1))[0]]
        print(f"\n--- Sheet: {sheet_name} ({ws.max_row} rows) ---")
        print(f"  Headers: {headers}")
        
        # Map column indices
        col_map = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            hl = h.lower()
            if 'uid' in hl or 'sr' in hl:
                col_map['uid'] = i
            if 'name' in hl and 'certified' in hl:
                col_map['name'] = i
            elif 'name' in hl and 'name' not in col_map:
                col_map['name'] = i
            if 'address' in hl or 'full address' in hl:
                col_map['address'] = i
            if 'long' in hl or 'lng' in hl:
                col_map['lng'] = i
            if 'lat' in hl:
                col_map['lat'] = i
            if 'special' in hl:
                col_map['specialisation'] = i
            if 'taluk' in hl:
                col_map['taluka'] = i
            if 'district' in hl:
                col_map['district'] = i
            if 'contact' in hl or 'phone' in hl:
                col_map['contact'] = i
            if 'pin' in hl:
                col_map['pincode'] = i
            if 'valid' in hl or 'date' in hl:
                col_map['validity'] = i
        
        print(f"  Column mapping: {col_map}")
        
        if 'name' not in col_map:
            print(f"  Skipping sheet (no name column found)")
            continue
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            
            name = vals[col_map['name']] if 'name' in col_map and col_map['name'] < len(vals) else None
            if not name or not str(name).strip():
                continue
            
            uid_val = vals[col_map['uid']] if 'uid' in col_map and col_map['uid'] < len(vals) else None
            uid = str(uid_val).strip() if uid_val else None
            
            lat = clean_coord(vals[col_map['lat']]) if 'lat' in col_map and col_map['lat'] < len(vals) else None
            lng = clean_coord(vals[col_map['lng']]) if 'lng' in col_map and col_map['lng'] < len(vals) else None
            
            address = vals[col_map['address']] if 'address' in col_map and col_map['address'] < len(vals) else None
            spec = vals[col_map['specialisation']] if 'specialisation' in col_map and col_map['specialisation'] < len(vals) else None
            taluka = vals[col_map['taluka']] if 'taluka' in col_map and col_map['taluka'] < len(vals) else None
            district = vals[col_map['district']] if 'district' in col_map and col_map['district'] < len(vals) else None
            contact = vals[col_map['contact']] if 'contact' in col_map and col_map['contact'] < len(vals) else None
            pincode = vals[col_map['pincode']] if 'pincode' in col_map and col_map['pincode'] < len(vals) else None
            validity = vals[col_map['validity']] if 'validity' in col_map and col_map['validity'] < len(vals) else None
            
            # Use name as dedup key since same healer may appear in multiple sheets
            key = str(name).strip()
            
            healer = all_healers.get(key, {})
            healer['name'] = str(name).strip()
            
            if uid:
                healer['uid'] = uid
            if lat is not None:
                healer['lat'] = lat
            if lng is not None:
                healer['lng'] = lng
            if address:
                healer['address'] = str(address).strip()
            if spec:
                # Merge specialisations
                existing_specs = set(s.strip() for s in healer.get('specialisation', '').split(',') if s.strip())
                new_specs = set(s.strip() for s in str(spec).split(',') if s.strip())
                all_specs = existing_specs | new_specs
                healer['specialisation'] = ', '.join(sorted(all_specs))
            if taluka:
                healer['taluka'] = str(taluka).strip()
            if district:
                healer['district'] = str(district).strip()
            if contact:
                healer['contact'] = str(int(contact) if isinstance(contact, float) else contact).strip()
            if pincode:
                healer['pincode'] = str(int(pincode) if isinstance(pincode, float) else pincode).strip()
            if validity:
                healer['validity'] = str(validity).strip()
            
            all_healers[key] = healer
    
    # Convert to list with IDs, filtering out entries without coordinates
    healers_list = []
    skipped = 0
    for i, (name, data) in enumerate(all_healers.items(), 1):
        if 'lat' not in data or 'lng' not in data or data.get('lat') is None or data.get('lng') is None:
            print(f"  WARNING: Skipping '{name}' - no coordinates")
            skipped += 1
            continue
        if 'specialisation' not in data or not data['specialisation']:
            data['specialisation'] = 'General'
        data['id'] = i
        healers_list.append(data)
    
    print(f"\n=== Results ===")
    print(f"Total unique healers found: {len(all_healers)}")
    print(f"Skipped (no coordinates): {skipped}")
    print(f"Valid healers to import: {len(healers_list)}")
    
    return healers_list


if __name__ == "__main__":
    healers = parse_excel()
    
    if healers:
        # Reassign sequential IDs
        for i, h in enumerate(healers, 1):
            h['id'] = i
        
        with open(HEALERS_JSON, "w", encoding="utf-8") as f:
            json.dump(healers, f, indent=2, ensure_ascii=False)
        
        print(f"\nWrote {len(healers)} healers to {HEALERS_JSON}")
        print("\nFirst 3 healers:")
        for h in healers[:3]:
            print(f"  {h['id']}: {h['name']} ({h.get('specialisation','?')}) @ {h['lat']},{h['lng']}")
    else:
        print("No valid healers found!")
